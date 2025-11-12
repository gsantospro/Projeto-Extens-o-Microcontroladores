import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# HELPERS PARA EXPORTAÇÃO EXCEL (SEM ALTERAÇÃO)
def _parse_hhmm(s):
    try:
        return datetime.strptime(s, "%H:%M").time()
    except Exception:
        return None

def calcular_horas_dia_excel(dia: dict) -> float:
    """
    Calcula horas do dia em FRAÇÃO DE DIA (padrão Excel):
    (saida - entrada) - (volta_intervalo - saida_intervalo)
    """
    e = _parse_hhmm(dia.get('entrada', ''))
    s = _parse_hhmm(dia.get('saida', ''))
    if not (e and s):
        return 0.0
    si = _parse_hhmm(dia.get('saida_intervalo', ''))
    vi = _parse_hhmm(dia.get('volta_intervalo', ''))

    base = (datetime.combine(datetime.today(), s) -
            datetime.combine(datetime.today(), e))
    if si and vi:
        base -= (datetime.combine(datetime.today(), vi) -
                 datetime.combine(datetime.today(), si))
    secs = max(0, base.total_seconds())
    return secs / 86400.0


def exportar_mes_xlsx(ano_mes: str, funcionarios: dict, registros: dict, eventos: list[str]) -> str:
    """
    Gera export/<MM-YYYY>_registros.xlsx
    - Uma aba por funcionário (com dados do mês)
    - Aba 'Resumo' com total de horas por funcionário
    Retorna caminho do arquivo gerado.
    """
    if len(ano_mes) != 7 or ano_mes[4] != "-":
        raise ValueError("Use o formato YYYY-MM (ex.: 2025-11)")

    os.makedirs("export", exist_ok=True)
    nome_arq = f"Registros_{datetime.strptime(ano_mes, '%Y-%m').strftime('%m-%Y')}.xlsx"
    caminho = os.path.join("export", nome_arq)

    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    totais_func = []
    created_any = False

    try:
        ano = int(ano_mes[:4])
        mes = int(ano_mes[5:])
        dt_inicio = datetime(ano, mes, 1)
        
        if mes == 12:
            dt_fim = datetime(ano + 1, 1, 1) - timedelta(days=1)
        else:
            dt_fim = datetime(ano, mes + 1, 1) - timedelta(days=1)
            
    except ValueError:
        raise ValueError("Formato de data YYYY-MM inválido.")


    for uid, nome in sorted(funcionarios.items(), key=lambda x: x[1].lower()):
        dias_mes = {d: dia for d, dia in (registros.get(uid, {}) or {}).items()
                    if d.startswith(ano_mes)}
        
        # Se não há registros, não precisa mais do 'continue'
        # if not dias_mes:
        #     continue

        created_any = True
        title_base = nome.strip().replace("/", "-").replace("\\", "-").replace(":", "-")
        title = title_base[:31] if title_base else uid[:8]
        ws = wb.create_sheet(title=title)

        cols = ["Data", "Dia", "Entrada", "Saída Intervalo", "Volta Intervalo", "Saída", "Horas"]
        ws.append(cols)
        for col_idx in range(1, len(cols)+1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center
            c.border = border_thin

        total_horas = 0.0
        
        current_date = dt_inicio
        while current_date <= dt_fim:
            data_str = current_date.strftime("%Y-%m-%d")
            
            dia = dias_mes.get(data_str, {})
            
            try:
                dt = current_date
                dia_semana = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"][dt.weekday()]
            except Exception:
                dia_semana = ""
                
            data_br = current_date.strftime("%d/%m/%Y")
            
            row = [
                data_br,
                dia_semana,
                dia.get("entrada", ""),
                dia.get("saida_intervalo", ""),
                dia.get("volta_intervalo", ""),
                dia.get("saida", ""),
                None,
            ]
            ws.append(row)

            r = ws.max_row
            horas_frac = calcular_horas_dia_excel(dia)
            total_horas += horas_frac
            c = ws.cell(row=r, column=7, value=horas_frac)
            c.number_format = "[h]:mm"
            c.alignment = right
            c.border = border_thin

            for col in range(1, 7):
                cell = ws.cell(row=r, column=col)
                cell.alignment = center
                cell.border = border_thin
                
            current_date += timedelta(days=1)

        ws.append(["", "", "", "", "TOTAL", "", total_horas])
        r = ws.max_row
        ws.cell(row=r, column=5).font = Font(bold=True)
        tc = ws.cell(row=r, column=7)
        tc.number_format = "[h]:mm"
        tc.font = Font(bold=True)
        tc.alignment = right
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = border_thin

        widths = [11, 6, 10, 16, 16, 10, 9]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:G{ws.max_row}"

        totais_func.append((nome, uid, total_horas))

    if "Sheet" in wb.sheetnames and not created_any:
        ws = wb["Sheet"]
        ws.title = "Resumo"
        ws.append(["Sem registros para", ano_mes])
    else:
        if "Sheet" in wb.sheetnames:
            std = wb["Sheet"]
            wb.remove(std)

    ws_r = wb.create_sheet("Resumo", 0)
    ws_r.append(["Funcionário", "UID", "Total Horas"])
    for col_idx in range(1, 4):
        c = ws_r.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border_thin

    for nome, uid, total in sorted(totais_func, key=lambda x: x[0].lower()):
        ws_r.append([nome, uid, total])
        r = ws_r.max_row
        ws_r.cell(row=r, column=3).number_format = "[h]:mm"
        for col in range(1, 4):
            ws_r.cell(row=r, column=col).border = border_thin
            ws_r.cell(row=r, column=col).alignment = center

    ws_r.column_dimensions["A"].width = 36
    ws_r.column_dimensions["B"].width = 20
    ws_r.column_dimensions["C"].width = 12
    ws_r.freeze_panes = "A2"
    ws_r.auto_filter.ref = f"A1:C{ws_r.max_row}"

    wb.save(caminho)
    return caminho